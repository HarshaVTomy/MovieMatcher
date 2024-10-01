import re
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, render,redirect
from django.contrib.auth.models import User
from django.contrib import auth
from django.contrib import messages
import requests
from .models import Customer, Movie, MovieList, Genre, Review
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.contrib.auth.password_validation import validate_password
from django.core.exceptions import ValidationError
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_str
from django.core.mail import EmailMultiAlternatives
from django.conf import settings
from django.http import HttpResponse
from django.contrib.auth.tokens import PasswordResetTokenGenerator
from django.template.loader import render_to_string

# Create your views here.
def index(request):
    movies = Movie.objects.all()

    # Safely get the last movie as featured
    if movies.exists():
        featured_movie = movies.last()  # Or adjust logic for selecting featured
    else:
        featured_movie = None

    context = {
        'movies': movies,
        'featured_movie': featured_movie  # Use singular for clarity
    }
    return render(request, 'index.html', context)

def customer_dashboard(request):
    return render(request, 'customer_dashboard.html')

@login_required
def edit_customer(request):
    # Retrieve the current logged-in user
    current_user = request.user
    # Check if the current user has a corresponding Customer instance
    try:
        customer = Customer.objects.get(user=current_user)
    except Customer.DoesNotExist:
        # Handle the case where the logged-in user does not have a corresponding Customer instance
        return HttpResponse("You are not associated with any customer profile.")

    if request.method == 'POST':
        # Update customer details with the data from the form
        customer.customer_name = request.POST['full-name']
        customer.email = request.POST['your-email']
        customer.contact_number = request.POST['phone-number']
        customer.save()

        # Update associated user's email
        current_user.email = request.POST['your-email']
        current_user.username = request.POST['your-email']
        current_user.save()

        # Redirect to the customer detail page after editing
        return redirect('/')

    # If it's a GET request, display the edit form with existing customer details
    return render(request, 'edit_customer.html', {'customer': customer})



@login_required(login_url='login')
def search(request):
    if request.method == 'POST':
        search_term = request.POST['search_term']
        movies = Movie.objects.filter(title__icontains=search_term)

        context = {
            'movies': movies,
            'search_term': search_term,
        }
        return render(request, 'search.html', context)
    else:
        return redirect('/')
    

def movies_by_genre(request, slug):
    genre = get_object_or_404(Genre, slug=slug)
    movies = genre.movies.all()  # Get movies in the genre
    return render(request, 'movies_by_genre.html', {'genre': genre, 'movies': movies})



@login_required(login_url='login')
def movie(request, pk):
    movie_uuid = pk
    movie_details = Movie.objects.get(uu_id=movie_uuid)

    context ={
        'movie_details': movie_details,
    }
    
    return render (request, 'movie.html', context)

# View for displaying movie details
def movie_detail(request, pk):
    movie = get_object_or_404(Movie, pk=pk)
    return render(request, 'movie.html', {'movie': movie})

# View to handle review submission
@login_required
def add_review(request, pk):
    movie = get_object_or_404(Movie, pk=pk)

    if request.method == 'POST':
        comment = request.POST.get('comment')
        rating = request.POST.get('rating')

        # Ensure the user doesn't post multiple reviews for the same movie
        if Review.objects.filter(movie=movie, user=request.user).exists():
            messages.error(request, "You have already reviewed this movie.")
        else:
            Review.objects.create(
                movie=movie,
                user=request.user,
                comment=comment,
                rating=rating
            )
            messages.success(request, "Your review has been submitted.")
        
    return redirect('movie_detail', pk=movie.pk)

@login_required(login_url='login')
def my_list(request):
    movie_list = MovieList.objects.filter(owner_user=request.user)
    user_movies = [movie.movie for movie in movie_list]

    # Get all genres from the user's list of movies
    user_genres = set()
    for movie in user_movies:
        user_genres.update(movie.genres.all())  # Assuming a many-to-many relationship for genres

    # Get similar movies excluding the ones already in the user's list
    similar_movies = Movie.objects.filter(genres__in=user_genres).exclude(id__in=[movie.id for movie in user_movies]).distinct()

    context = {
        'movies': user_movies,
        'similar_movies': similar_movies,  # Add this line
    }
    return render(request, 'my_list.html', context)


@login_required(login_url='login')
def add_to_list(request):
    if request.method == 'POST':
        movie_url_id = request.POST.get('movie_id')
        uuid_pattern = r'[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}'
        match = re.search(uuid_pattern, movie_url_id)
        movie_id = match.group() if match else None

        movie = get_object_or_404(Movie, uu_id=movie_id)
        movie_list, created = MovieList.objects.get_or_create(owner_user=request.user, movie=movie)

        if created:
            response_data = {'status': 'success', 'message': 'Added ✓'}
        else:
            response_data = {'status': 'info', 'message': 'Movie already in list'}

        return JsonResponse(response_data)
    else:
        # return error
        return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=400)

def remove_from_list(request):
    if request.method == "POST" and request.user.is_authenticated:
        movie_id = request.POST.get('movie_id')  # Get the movie ID from the POST request
        movie = get_object_or_404(Movie, id=movie_id)  # Fetch the movie object
        
        # Find the MovieList entry for the current user and selected movie
        movie_list_item = MovieList.objects.filter(owner_user=request.user, movie=movie)

        if movie_list_item.exists():
            movie_list_item.delete()  # Remove the movie from the user's list
            return JsonResponse({'message': 'Movie removed from your list.'}, status=200)
        else:
            return JsonResponse({'message': 'Movie not found in your list.'}, status=404)

    return JsonResponse({'message': 'Invalid request.'}, status=400)


def register(request):
    if request.method == 'POST':
        full_name = request.POST.get('full-name')
        email = request.POST.get('your-email')
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm-password')
        phone_number = request.POST.get('phone-number')

        # Check if passwords match
        if password != confirm_password:
            messages.error(request, 'Passwords Do Not Match!')
            return render(request, 'register.html')

        # Check password strength using Django's built-in validators
        try:
            validate_password(password)
        except ValidationError as e:
            messages.error(request, ', '.join(e.messages))
            return render(request, 'register.html')

        # Create user
        try:
            user = User.objects.create_user(username=email, email=email, password=password)
        except:
            messages.error(request, 'Email already exists!')
            return render(request, 'register.html')

        # Create customer
        customer = Customer.objects.create(user=user, customer_name=full_name, email=email, contact_number=phone_number)
        
        # Authenticate and login user
        user = authenticate(request, username=email, password=password)
        if user is not None:
            login(request, user)
            messages.success(request, 'Your Account Has Been Registered Successfully!')
            return redirect('index')
        else:
            messages.error(request, 'Failed to login user.')
            return render(request, 'register.html')

    return render(request, 'register.html')



def login_view(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')
        
        user = authenticate(request, username=email, password=password)
        
        if user is not None:
            login(request, user)
            messages.success(request, 'You have successfully logged in!')
            return redirect('index')  # Change 'home' to your desired redirect URL after login
        else:
            messages.error(request, 'Invalid email or password. Please try again.')
            return render(request, 'login.html')  # Change 'customer/login.html' to your login template path
    return render(request, 'login.html')  # Change 'login.html' to your login template path

def user_logout(request):
    logout(request)
    return redirect('index') 



def change_password(request):
    if request.method == 'POST':
        old_password = request.POST.get('old_password')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')
        user = request.user  # Assuming the user is already logged in

        if user.check_password(old_password):
            if new_password == confirm_password:
                user.set_password(new_password)
                user.save()
                messages.success(request, "Password changed successfully.")
                return redirect('login')
            else:
                messages.error(request, "New passwords do not match.")
        else:
            messages.error(request, "Old password is incorrect.")

    return render(request, 'change_password.html')

class CustomTokenGenerator(PasswordResetTokenGenerator):
    def _make_hash_value(self, user, timestamp):
        return (
            str(user.pk) + user.password + str(timestamp)
        )



def forgot_password(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        try:
            user = User.objects.get(email=email)
            uid = urlsafe_base64_encode(force_bytes(user.pk))
            token = CustomTokenGenerator().make_token(user)
            reset_password_url = request.build_absolute_uri('/reset_password/{}/{}/'.format(uid, token))
            email_subject = 'Reset Your Password'

            # Render both HTML and plain text versions of the email
            email_body_html = render_to_string('reset_password_email.html', {
                'reset_password_url': reset_password_url,
                'user': user,
            })
            email_body_text = "Click the following link to reset your password: {}".format(reset_password_url)

            # Create an EmailMultiAlternatives object to send both HTML and plain text versions
            email = EmailMultiAlternatives(
                email_subject,
                email_body_text,
                settings.EMAIL_HOST_USER,
                [email],
            )
            email.attach_alternative(email_body_html, 'text/html')  # Attach HTML version
            email.send(fail_silently=False)

            messages.success(request, 'An email has been sent to your email address with instructions on how to reset your password.')
            return redirect('login')
        except User.DoesNotExist:
            messages.error(request, "User with this email does not exist.")
    return render(request, 'forgot_password.html')


def reset_password(request, uidb64, token):
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        user = None

    if user is not None and CustomTokenGenerator().check_token(user, token):
        if request.method == 'POST':
            new_password = request.POST.get('new_password')
            confirm_password = request.POST.get('confirm_password')

            if new_password == confirm_password:
                user.set_password(new_password)
                user.save()
                messages.success(request, "Password reset successfully. You can now login with your new password.")
                return redirect('login')
            else:
                messages.error(request, "Passwords do not match.")
        return render(request, 'reset_password.html')
    else:
        messages.error(request, "Invalid reset link. Please try again or request a new reset link.")
        return redirect('login')


import openpyxl
from django.http import HttpResponse
from .models import Movie

def top_movies_report(request):
    # Query the top 5 movies based on IMDb rating
    top_movies = Movie.objects.all().order_by('-imdb_rating')[:5]

    # Check if the user requested an export as Excel
    if request.GET.get('export') == 'true':
        # Create an in-memory workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Top Movies"

        # Define headers
        headers = ['Title', 'Director', 'IMDb Rating', 'Release Date']
        ws.append(headers)

        # Append movie data
        for movie in top_movies:
            ws.append([movie.title, movie.movie_director, movie.imdb_rating, movie.release_date])

        # Prepare response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=top_5_movies.xlsx'

        # Save the workbook to the response
        wb.save(response)
        return response

    # Otherwise, render the HTML page
    return render(request, 'top_movies_report.html', {'top_movies': top_movies})



from django.shortcuts import render
import google.generativeai as genai

# Configure Gemini API with your API key
genai.configure(api_key="AIzaSyDjhqNBjd_zqrkvlSI2_ELJddEEr9SPTS8")

def movie_recommendation(request):
    recommendation = None
    
    if request.method == "POST":
        rating = request.POST['rating']  # Expecting a string like "4-5"
        genre = request.POST['genre']
        release_year = request.POST['release_year']
        actor = request.POST['actor']
        mood = request.POST['mood']

        # Define the prompt to send to the Gemini API
        user_input = f"""
        I'm looking for movies that match the following criteria:
        - Genre: {genre}
        - Release Year: {release_year}
        - Lead Actor: {actor}
        - Rating: {rating}
        - Mood: {mood}
        
        Can you suggest 2-3 movies and provide a brief description of each?
        """

        # Call the Gemini model
        model = genai.GenerativeModel(model_name="gemini-1.5-flash")
        chat_session = model.start_chat(history=[])
        response = chat_session.send_message(user_input)
        
        recommendation = response.text  # Fetch the recommendation result

    return render(request, 'movie_recommendation.html', {'recommendation': recommendation})



# Replace with your API key
API_KEY = 'AIzaSyDWn7FhpBcLeXw4ygXZd-86c-es1LJiWBA'
API_URL = 'https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash-latest:generateContent'

# Define predefined responses
PREDEFINED_RESPONSES = {
    "Can you tell me what this THRIVEseeds is about?": "Hi, I’m Cropsy! Our e-commerce website THRIVEseeds specializes in selling high-quality crop seeds for various agricultural needs. We offer a wide range of seeds with detailed descriptions, pricing, and weather-based recommendations.",
    "What kinds of crop seeds do you sell?": "Cropsy here! We offer a diverse range of crop seeds including vegetables, fruits, grains, and pulses. You can browse our categories to find specific types of seeds.",
    "Can you give me details about a specific seed?": "Sure thing! Just tell me the name or category of the seed you’re interested in, and I’ll provide you with more details.",
    "How does weather affect the seeds I should buy?": "Great question! Weather plays a crucial role in crop growth. I can help you choose the right seeds based on your local climate conditions using our weather forecasts.",
    "Can you recommend seeds based on the current weather?": "Absolutely! Based on your location and the current weather conditions, I can suggest the best seeds for optimal growth. Just let me know your location.",
    "How do I add items to my cart?": "To add items to your cart, simply select the desired seed, choose the quantity, and click the 'Add to Cart' button.",
    "I want to remove an item from my cart. How do I do that?": "No problem! Go to your cart page, find the item you want to remove, and click the 'Remove' button next to it.",
    "How do I check out?": "To check out, go to your cart, review the items, and click the 'Proceed to Checkout' button. Follow the prompts to enter your shipping information and payment details.",
    "I have a problem with my order. Who should I contact?": "If you have any issues with your order, please contact our customer support team through the contact form on our website or by email at harshavtomy01@gmail.com.",
    "How can I track my order?": "You can track your order by visiting the 'Order Tracking' section on our website and entering your order number.",
    "How do I create an account?": "To create an account, click on the 'Sign Up' button on the homepage, fill out the required information, and submit the form. You’ll receive a confirmation email to complete the registration.",
    "How can I reset my password?": "If you’ve forgotten your password, go to the 'Login' page and click on 'Forgot Password.' Follow the instructions to reset your password.",
    "What weather conditions should I consider when buying seeds?": "When purchasing seeds, you should consider factors like temperature, humidity, rainfall, and soil conditions. I can provide you with weather forecasts to help you make the right decision.",
    "Can you give me today’s weather forecast?": "Sure! Let me check the current weather conditions for your location. Could you share your city or town?",
    "What is the weather forecast for the next 7 days?": "I can provide you with a 7-day weather forecast for your area. Please visit \"weather dashboard\" after login for get weather forecasting data up to 16 days from now.",
    "How does the weather forecasting feature work?": "THRIVEseeds integrates weather data from reliable sources to help you make informed decisions. The forecasts are updated regularly, and I can provide real-time information for your specific area.",
    "What are the available payment options?": "We accept major credit cards, debit cards, UPI, and net banking. You can choose your preferred option during checkout.",
    "How can I contact customer support?": "You can contact our customer support through the contact form on our website or by emailing us at harshavtomy01@gmail.com.",
    "What’s your favorite color?": "As much as I’d love to have a favorite color, I’m here to help you with crop seed-related queries! Let me know if you need assistance with any products or weather updates.",
    "Tell me a joke.": "I’m more of a seed and weather expert, but I can certainly help you grow some great crops! Let me know if you need assistance with anything else.",
    "How do I fix my car engine?": "I specialize in crop seeds and weather forecasting, so I might not be able to help with that. However, if you have any questions about our products, I’d be happy to assist!",
    "Can you predict the stock market for me?": "I’m here to provide you with weather forecasts and help with crop seed-related queries. If you’re looking for investment advice, I recommend contacting a financial expert.",
    "How can I grow flowers in space?": "That’s an exciting question! While I can help you grow crops on Earth, space gardening is a bit out of my expertise. Let me know if you need any tips on planting crops here on Earth.",
    "Can you recommend seeds based on the current weather?": "Absolutely! To recommend the best seeds for your area, I need to know your location. Please tell me your city or town.",
    "how are you": "I'm just a chatbot here to assist you with crop seed-related questions. How can I help you today?",
}



def chat(request):
    if request.method == 'POST':
        user_message = request.POST.get('message')

        # Retrieve or initialize conversation history
        conversation_history = request.session.get('conversation_history', [])

        # Add user message to conversation history
        conversation_history.append(f"input: {user_message}")

        # Check if the message matches any predefined response
        bot_reply = PREDEFINED_RESPONSES.get(user_message, None)
        
        if not bot_reply:
            # Define headers and data for the API request
            headers = {
                'Content-Type': 'application/json',
            }
            
            # Prepare context: Use the conversation history
            messages = [{'text': message} for message in conversation_history]
            
            # Prepare data with context (previous conversation)
            data = {
                'contents': [
                    {
                        'parts': messages
                    }
                ]
            }

            # Make the API request
            try:
                response = requests.post(f'{API_URL}?key={API_KEY}', headers=headers, json=data)
                response.raise_for_status()  # Raise an exception for HTTP errors
                
                # Parse the JSON response
                api_response = response.json()
                print("API Response:", api_response)  # For debugging
                
                # Extract the bot reply from the response
                bot_reply = api_response['candidates'][0]['content']['parts'][0]['text']
                
                # Limit the response to a certain number of sentences (e.g., 3)
                bot_reply = '. '.join(bot_reply.split('. ')[:3])  # Limits the response to 3 sentences
                
            except requests.RequestException as e:
                # Handle request errors
                print(f"API request error: {e}")
                bot_reply = 'Sorry, there was an error processing your request.'

        # Add bot response to conversation history
        conversation_history.append(f"output: {bot_reply}")

        # Store updated conversation history in session
        request.session['conversation_history'] = conversation_history

        return JsonResponse({'reply': bot_reply})

    # Render the chat interface if not a POST request
    return render(request, 'chatbot.html')

import json
from django.db.models import Count, Sum, Avg
from django.core.serializers.json import DjangoJSONEncoder


def visualization_view(request):
    # Number of movies per genre
    genre_data = Genre.objects.annotate(movie_count=Count('movies')).values('name', 'movie_count')
    
    # IMDb rating distribution
    imdb_rating_data = Movie.objects.values('imdb_rating').annotate(count=Count('id')).order_by('imdb_rating')

    # Convert QuerySets to JSON serializable lists
    genre_data = json.dumps(list(genre_data), cls=DjangoJSONEncoder)
    imdb_rating_data = json.dumps(list(imdb_rating_data), cls=DjangoJSONEncoder)

    # Pass data to the template
    context = {
        'genre_data': genre_data,
        'imdb_rating_data': imdb_rating_data
    }

    return render(request, 'visualizations.html', context)

from django.db.models import Sum, F
from django.http import HttpResponse
from django.shortcuts import render
from openpyxl.utils import get_column_letter
import openpyxl
from datetime import datetime

def customer_report(request):
    # Get all customers from the database
    customers = Customer.objects.all()

    if 'export' in request.GET:
        # Create an Excel workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Customer Report"

        # Write the headers
        headers = ['Name', 'Email', 'Contact Number', 'Role']
        for col_num, header in enumerate(headers, 1):
            ws[f"{get_column_letter(col_num)}1"] = header

        # Write data to the Excel sheet
        for row_num, customer in enumerate(customers, 2):
            ws[f"A{row_num}"] = customer.customer_name
            ws[f"B{row_num}"] = customer.email
            ws[f"C{row_num}"] = customer.contact_number
            ws[f"D{row_num}"] = customer.role

        # Prepare the response to download the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Customer_Report.xlsx'
        wb.save(response)
        return response

    # Context for rendering the HTML page
    context = {
        'customers': customers,
    }
    return render(request, 'customer_report.html', context)

def about(request):
    return render(request, 'about.html')

def privacy_policy(request):
    return render(request, 'privacy-policy.html')

def terms(request):
    return render(request, 'terms.html')
